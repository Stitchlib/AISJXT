import json
import os
import tempfile
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import FileResponse

from ..auth import get_current_user
from ..models import ReportSummary, TrendPoint, TypeShare, User

router = APIRouter(prefix="/reports", tags=["reports"], dependencies=[Depends(get_current_user)])


@router.get("/summary", response_model=ReportSummary)
def summary(request: Request, bucket: str = "day"):
    db = request.app.state.db
    stats = db.get_statistics()
    shares = [TypeShare(**s) for s in db.get_type_shares()]
    trend = [TrendPoint(**t) for t in db.get_trend(bucket)]
    return ReportSummary(
        total=stats["total"],
        defect_count=stats["defect_count"],
        defect_rate=stats["defect_rate"],
        avg_processing_ms=stats["avg_processing_ms"],
        by_type=shares,
        trend=trend,
    )


@router.get("/export")
def export(request: Request, format: str = "excel"):
    db = request.app.state.db
    tmp = tempfile.gettempdir()
    if format == "csv":
        path = db.export_csv(os.path.join(tmp, "aiqc_report.csv"))
        return FileResponse(path, filename="report.csv", media_type="text/csv")
    if format != "excel":
        raise HTTPException(status_code=400, detail="仅支持 excel / csv")

    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl")

    stats = db.get_statistics()
    shares = db.get_type_shares()
    trend = db.get_trend("day")
    rows, _ = db.query_results(1, 100_000)

    wb = Workbook()
    ws = wb.active
    ws.title = "概览"
    ws.append(["指标", "值"])
    for label, val in [
        ("检测总数", stats["total"]),
        ("缺陷总数", stats["defect_count"]),
        ("不良率", stats["defect_rate"]),
        ("平均处理耗时(ms)", stats["avg_processing_ms"]),
        ("仿真记录数", stats["simulated_records"]),
    ]:
        ws.append([label, val])

    ws2 = wb.create_sheet("瑕疵类型占比")
    ws2.append(["瑕疵类型", "数量"])
    for s in shares:
        ws2.append([s["class_name"], s["count"]])

    ws3 = wb.create_sheet("不良率趋势(按日)")
    ws3.append(["日期", "检测数", "缺陷数", "不良率"])
    for t in trend:
        ws3.append([t["bucket"], t["total"], t["defect_count"], t["defect_rate"]])

    ws4 = wb.create_sheet("明细")
    ws4.append(["ID", "时间", "摄像头", "缺陷数", "检测数", "不良率", "耗时ms", "仿真", "瑕疵明细"])
    for r in rows:
        ws4.append([
            r["id"], r["timestamp"], r["camera_id"], r["defect_count"], r["total_count"],
            r["defect_rate"], r["processing_time_ms"], r["is_simulation"],
            json.dumps(r.get("defects", []), ensure_ascii=False),
        ])

    path = os.path.join(tmp, "aiqc_report.xlsx")
    wb.save(path)
    return FileResponse(
        path, filename="quality_report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
